import openpyxl
from urlToApiIP import converter

#For converting URL list of excel files to Api IP excel files

def newWriter():
    path = input("Enter the path to the file: ")
    print("File Path is: " + path)
    wb = openpyxl.load_workbook(path)
    wb2 = openpyxl.Workbook()
    writeSheet = wb2.active
    sheetNames = wb.get_sheet_names()
    for i in sheetNames:
        sheet = wb[i]
        for j in range(2, sheet.max_row + 1):
            user, ID = converter(sheet.cell(row = j, column = 2).value)
            row = [user, ID]
            writeSheet.append(row)
    wb2.save("IP.xlsx")

if __name__ == "__main__":
    newWriter()