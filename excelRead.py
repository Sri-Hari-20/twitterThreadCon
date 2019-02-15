import excelWrite
import openpyxl

def excelReader(path):
    wb = openpyxl.load_workbook(path)
    workSheet = wb.active
    maxRows = workSheet.max_row
    for i in range(2, maxRows):
        url = workSheet.cell(row = i, column = 2).value
        excelWrite.writeMode(path, url)

if __name__ == "__main__":
    path = input("Enter the file path to generate the replies for: ")
    excelReader(path)
